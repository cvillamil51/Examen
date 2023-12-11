from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time 
#Inicializar el WebDriver para Firefox
s = Service(GeckoDriverManager().install())
driver = webdriver.Firefox(service=s) 

# Abrir una página web
driver.get("https://portal-test.zoefin.com/reschedule/66965aa0-9c17-11ed-b52a-53af9ee871d5")

# Espera hasta que el elemento sea visible
element = WebDriverWait(driver, 10).until(
    EC.visibility_of_element_located((By.XPATH, "//div[@class='styles-module_datesContainer__sZsX5 styles-module_datesContainer__modActive__X5cq1 styles-module_datesContainer__modIsDesktop__U6mIw']"))
)

# Encuentra el elemento Type.
Dia = driver.find_element(By.XPATH, "//div[@data-testid='calendar-day-11']//span[contains(text(),'11')]")
# Dar clic en los elementos.
Dia.click()
# Espera hasta que el elemento sea visible
element = WebDriverWait(driver, 10).until(
    EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'12:30 pm')]"))
)
Hora = driver.find_element(By.XPATH, "//div[contains(text(),'12:30 pm')]")
# Dar clic en los elementos.
Hora.click()
forma = driver.find_element(By.XPATH, "//p[contains(text(),'Video Conference')]")
button = driver.find_element(By.XPATH, "//button[contains(text(),'reschedule')]")
# Dar clic en los elementos.
forma.click()
button.click()
# Validacion si existe el elemento sea visible poUp Confirm reschedule
try:
    # Espera hasta 10 segundos para que aparezca el elemento "//div[@class='sc-aXZVg iCWQkn']"
    elemento = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='sc-aXZVg iCWQkn']"))
    )
    # Verifica si el elemento con el XPath dado está presente
    if  driver.find_elements(By.XPATH,"//div[@class='sc-aXZVg iCWQkn']"):
        # Abre el archivo Excel
        workbook = load_workbook(r'C:\\Users\\LUBISEX\\Downloads\\Examen\\Test_Suite.xlsx')
        sheet = workbook.active

        # Ingresa "X" en la celda E8 de la hoja 1 (fila 8, columna 5)
        sheet.cell(row=8, column=5, value="X")

        # Guarda el archivo Excel y ciérralo
        workbook.save(r'C:\\Users\\LUBISEX\\Downloads\\Examen\\Test_Suite.xlsx')
        workbook.close()
    else:
        excel_file = (r'C:\\Users\\LUBISEX\\Downloads\\Examen\\Test_Suite.xlsx')
        # Abre el archivo Excel
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        # Ingresa "X" en la celda F6 de la hoja 1 (fila 7, columna 5)
        sheet.cell(row=8, column=6, value="X")
        # Guarda el archivo Excel y ciérralo
        workbook.save(excel_file)
        workbook.close()
        # Toma la captura de pantalla y guárdala en la ruta especificada
        ruta_captura = r"'C:\\Users\\LUBISEX\\Downloads\\Examen\\Evidencia\\EvidenciaPopUp.png'"
        driver.save_screenshot(ruta_captura)
        print(f"Captura de pantalla guardada en: {ruta_captura}")
finally:
# Tiempo de espera de 4 segundos
    time.sleep(7)
    driver.quit()
